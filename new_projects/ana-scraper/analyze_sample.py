"""
analyze_sample.py  - reads the client's sample CSV and print columns + first 3 rows
also prints Excel columns I and biweekly structure
"""
import csv, sys

# CSV
with open('config/サンプル_抽出結果ファイル.csv', encoding='utf-8-sig', errors='replace') as f:
    reader = csv.reader(f)
    rows = list(reader)

print('=== SAMPLE CSV ===')
print('Total rows (inc header):', len(rows))
if rows:
    print('Columns:', rows[0])
    for i,r in enumerate(rows[1:4], 1):
        print(f'Row {i}:', r)

# Excel
import openpyxl
wb = openpyxl.load_workbook('config/◆◆価格比較対象リスト(曜日別).xlsx', data_only=True)
ws = wb.active
print('\n=== EXCEL ===')
print('Sheet:', ws.title, ' rows:', ws.max_row)
print('Row 21 (header):')
h = list(ws.iter_rows(min_row=21, max_row=21, values_only=True))[0]
for i, v in enumerate(h):
    if v: print(f'  col[{i}] = {v!r}')

print('\nRows 22-24 (data):')
for rownum, row in enumerate(ws.iter_rows(min_row=22, max_row=24, values_only=True), 22):
    print(f'  row {rownum}:', row)

# Check column I (index 8) for biweekly/150 values
print('\nColumn I (index 8) distinct values:')
vals = set()
for row in ws.iter_rows(min_row=22, values_only=True):
    v = row[8]
    if v: vals.add(str(v))
print(' ', vals)
