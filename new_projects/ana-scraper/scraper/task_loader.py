"""
Task Loader — reads 255+ search task definitions from the client's Excel file.

Excel structure (sheet: 【最新】 20251009):
  - Header row: Row 21
  - Data rows:  Row 22 onwards

Column mapping (0-indexed):
  Col 2  → task_id         (sequential int)
  Col 3  → run_weekday     ('Sat', 'Sun', 'Mon', 'Tue', 'Wed', 'Thu', 'Fri')
  Col 5  → jpe_jpw         ('JPE' | 'JPW')
  Col 7  → destination     (area name in Japanese, e.g. '沖縄', '関東', '九州')
  Col 10 → pattern         (description string, e.g. '羽田⇒石垣 高シェア')
  Col 11 → dept_airport    (IATA code, e.g. 'HND', 'ITM')
  Col 12 → arr_airport     (IATA code, e.g. 'ISG', 'OKA')
  Col 13 → go_flight_filter ('ANA' = match ANA flights | '最安値' = cheapest, any)
  Col 14 → go_flight_num   (int ANA flight number suffix, e.g. 89 → NH089)
  Col 15 → rt_flight_filter
  Col 16 → rt_flight_num   (int)
  Col 17 → nights          (int, 1 / 2 / 3)
  Col 18 → adults          (int, 1 or 2)
  Col 19 → hotel_name      (Target hotel name for post-filter, may be None)
  Col 20 → flag            (1 = active)

district_code is derived from destination area.
date_range_days defaults to 90.
"""

import logging
import glob
from dataclasses import dataclass, field
from pathlib import Path
from typing import List, Optional

logger = logging.getLogger(__name__)


# ------------------------------------------------------------------
# District code mapping (destination area → ANA districtCd)
# ------------------------------------------------------------------
DESTINATION_TO_DISTRICT: dict[str, str] = {
    "北海道": "02",
    "東北": "03",
    "仙台": "03",
    "関東": "04",
    "関東・甲信越": "04",
    "東京": "04",
    "東海": "05",
    "北陸": "05",
    "中部": "05",        # Tokai/Chubu (Nagoya area)
    "名古屋": "05",
    "近畿": "06",
    "関西": "06",
    "大阪": "06",
    "中国": "07",
    "四国": "07",
    "中四国": "07",      # Chugoku + Shikoku combined
    "広島": "07",
    "九州": "08",
    "沖縄": "08",
    "九州・沖縄": "08",
    "福岡": "08",
    "鹿児島": "08",
}


# Weekday string → Python weekday number (0=Mon, 6=Sun)
WEEKDAY_MAP: dict[str, int] = {
    "Mon": 0,
    "Tue": 1,
    "Wed": 2,
    "Thu": 3,
    "Fri": 4,
    "Sat": 5,
    "Sun": 6,
}


@dataclass
class SearchTask:
    task_id: int
    dept_airport: str           # e.g. "HND"
    arr_airport: str            # e.g. "ISG"
    district_code: str          # e.g. "08"
    destination: str = ""       # e.g. "沖縄"
    pattern: str = ""           # e.g. "羽田⇒石垣 高シェア"
    region_code: str = ""
    adults: int = 2
    nights: int = 2
    date_range_days: int = 90   # 90 (default) or 150 (from col I)
    is_biweekly: bool = False   # True when col I contains '隔週'
    run_weekday: int = 5        # 0=Mon … 6=Sun
    hotel_name: Optional[str] = None
    go_flight_filter: str = "最安値"    # "ANA" | "最安値"
    go_flight_num: Optional[int] = None # numeric suffix e.g. 89 → NH089
    rt_flight_filter: str = "最安値"
    rt_flight_num: Optional[int] = None
    is_high_share: bool = False  # True if pattern contains "高シェア"
    jpe_jpw: str = ""


def _derive_district(destination: str) -> str:
    """Map a Japanese destination area name to ANA district code."""
    if not destination:
        return "08"
    for key, code in DESTINATION_TO_DISTRICT.items():
        if key in destination:
            return code
    logger.warning("Unknown destination '%s', defaulting district_code='08'", destination)
    return "08"


def _parse_weekday(val) -> int:
    """Convert weekday string or int to Python weekday number."""
    if isinstance(val, int):
        return val % 7
    if isinstance(val, str):
        return WEEKDAY_MAP.get(val.strip()[:3], 5)
    return 5  # default Saturday


def _format_flight_no(prefix: str, number: Optional[int]) -> Optional[str]:
    """Format flight label like 'NH089' when prefix='ANA' and number=89."""
    if prefix != "ANA" or number is None:
        return None
    return f"NH{int(number):04d}"


# ------------------------------------------------------------------
# Fallback sample tasks for testing
# ------------------------------------------------------------------
SAMPLE_TASKS: List[SearchTask] = [
    SearchTask(
        task_id=1,
        dept_airport="HND",
        arr_airport="ISG",
        district_code="08",
        destination="沖縄",
        pattern="羽田⇒石垣 高シェア",
        adults=2,
        nights=2,
        date_range_days=90,
        run_weekday=5,  # Saturday
        hotel_name="アートホテル石垣島",
        go_flight_filter="ANA",
        go_flight_num=89,
        rt_flight_filter="ANA",
        rt_flight_num=92,
        is_high_share=True,
        jpe_jpw="JPW",
    ),
    SearchTask(
        task_id=2,
        dept_airport="HND",
        arr_airport="ISG",
        district_code="08",
        destination="沖縄",
        pattern="羽田⇒石垣 最安値",
        adults=2,
        nights=2,
        date_range_days=90,
        run_weekday=5,
        hotel_name="アートホテル石垣島",
        go_flight_filter="最安値",
        rt_flight_filter="最安値",
        is_high_share=False,
        jpe_jpw="JPW",
    ),
]


def load_tasks(excel_path: Optional[str] = None) -> List[SearchTask]:
    """
    Load tasks from the client's Excel file.

    Auto-discovers the Excel file in the config/ directory if the path
    given is the old placeholder 'tasks.xlsx' and that file doesn't exist.

    Falls back to SAMPLE_TASKS if the file cannot be found/parsed.
    """
    # Auto-discover the actual Excel file
    if excel_path:
        path = Path(excel_path)
        if not path.exists():
            # Try to auto-discover *.xlsx in the same directory
            parent = path.parent if path.parent != Path(".") else Path("config")
            candidates = list(parent.glob("*.xlsx"))
            if candidates:
                path = candidates[0]
                logger.info("Auto-discovered Excel: %s", path)
            else:
                logger.warning("Excel file not found: %s — using SAMPLE_TASKS", excel_path)
                return SAMPLE_TASKS
    else:
        # Try config/ directory
        candidates = list(Path("config").glob("*.xlsx"))
        if not candidates:
            logger.warning("No Excel file found in config/ — using SAMPLE_TASKS")
            return SAMPLE_TASKS
        path = candidates[0]
        logger.info("Auto-discovered Excel: %s", path)

    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(path), data_only=True)
        ws = wb.active
        logger.info("Opened Excel: %s / sheet: %s (rows=%d)", path, ws.title, ws.max_row)

        tasks: List[SearchTask] = []
        skipped = 0

        # Data starts at row 22 (header is row 21)
        for row in ws.iter_rows(min_row=22, values_only=True):
            task_id_raw = row[2]  # Col C

            # Skip rows without a numeric task_id
            if task_id_raw is None or not isinstance(task_id_raw, (int, float)):
                skipped += 1
                continue

            task_id = int(task_id_raw)
            flag = row[20]  # Col U — skip inactive tasks (flag != 1)
            if flag is not None and flag != 1:
                skipped += 1
                continue

            run_weekday_raw = row[3]    # Col D
            exec_week_raw  = str(row[8] or "").strip()   # Col I  ('隔週' | '150' | '')
            destination = str(row[7] or "").strip()   # Col H
            pattern = str(row[10] or "").strip()       # Col K
            dept_airport = str(row[11] or "").strip()  # Col L
            arr_airport = str(row[12] or "").strip()   # Col M
            go_filter = str(row[13] or "最安値").strip()  # Col N
            go_num_raw = row[14]   # Col O
            rt_filter = str(row[15] or "最安値").strip()  # Col P
            rt_num_raw = row[16]   # Col Q
            nights_raw = row[17]   # Col R
            adults_raw = row[18]   # Col S
            hotel_name = str(row[19]).strip() if row[19] else None  # Col T
            jpe_jpw = str(row[5] or "").strip()  # Col F

            # Validate required fields
            if not dept_airport or not arr_airport:
                skipped += 1
                continue

            go_flight_num = int(go_num_raw) if isinstance(go_num_raw, (int, float)) else None
            rt_flight_num = int(rt_num_raw) if isinstance(rt_num_raw, (int, float)) else None
            nights = int(nights_raw) if isinstance(nights_raw, (int, float)) else 2
            adults = int(adults_raw) if isinstance(adults_raw, (int, float)) else 2

            # Col I: '150' → 150-day fetch; '隔週' → biweekly run; both can coexist
            is_biweekly     = "隔週" in exec_week_raw
            date_range_days = 150 if "150" in exec_week_raw else 90

            task = SearchTask(
                task_id=task_id,
                dept_airport=dept_airport,
                arr_airport=arr_airport,
                district_code=_derive_district(destination),
                destination=destination,
                pattern=pattern,
                adults=adults,
                nights=nights,
                date_range_days=date_range_days,
                is_biweekly=is_biweekly,
                run_weekday=_parse_weekday(run_weekday_raw),
                hotel_name=hotel_name,
                go_flight_filter=go_filter,
                go_flight_num=go_flight_num,
                rt_flight_filter=rt_filter,
                rt_flight_num=rt_flight_num,
                is_high_share="高シェア" in pattern,
                jpe_jpw=jpe_jpw,
            )
            tasks.append(task)

        logger.info(
            "Loaded %d tasks from %s (skipped %d rows)", len(tasks), path, skipped
        )
        return tasks

    except ImportError:
        logger.error("openpyxl not installed — pip install openpyxl")
        return SAMPLE_TASKS
    except Exception as e:
        logger.error("Failed to load Excel '%s': %s — using SAMPLE_TASKS", path, e)
        return SAMPLE_TASKS


def get_go_flight_label(task: SearchTask) -> Optional[str]:
    """Return ANA flight label like 'NH0089' for filtering, or None for cheapest."""
    return _format_flight_no(task.go_flight_filter, task.go_flight_num)


def get_rt_flight_label(task: SearchTask) -> Optional[str]:
    """Return ANA return flight label like 'NH0092' for filtering, or None for cheapest."""
    return _format_flight_no(task.rt_flight_filter, task.rt_flight_num)
